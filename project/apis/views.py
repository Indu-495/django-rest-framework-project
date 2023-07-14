import io
import os
from django.shortcuts import render

# Create your views here.
from django.shortcuts import render
import requests
from .models import *
from .serializers import *
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.generics import GenericAPIView
from rest_framework.mixins import ListModelMixin,CreateModelMixin,RetrieveModelMixin,UpdateModelMixin,DestroyModelMixin
from rest_framework.generics import ListAPIView,CreateAPIView,RetrieveAPIView,UpdateAPIView,DestroyAPIView,ListCreateAPIView,RetrieveUpdateAPIView,RetrieveDestroyAPIView,RetrieveUpdateDestroyAPIView
from rest_framework import viewsets
from rest_framework.authentication import BasicAuthentication,SessionAuthentication
from rest_framework.permissions import IsAuthenticated, IsAdminUser

# views.py

from django.http import FileResponse, HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from reportlab.pdfgen import canvas
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from django.http import HttpResponse
from django.shortcuts import render
from datetime import datetime
from .models import CallReportMaster
from django.shortcuts import render

def create_html(request):
    if request.method == 'POST':
        # Handle the form submission and API interaction
        emp_id = request.POST.get('emp_id')
        name = request.POST.get('name')
        month = request.POST.get('month')
        pf_status = request.POST.get('pf_status')
        gross = request.POST.get('gross')
        net_days = request.POST.get('net_days')
        arrears = request.POST.get('arrears')
        shift_allowance = request.POST.get('shift_allowance')

        # Perform any necessary validation or data processing

        # Call the API endpoint to create the employee
        api_url = 'http://127.0.0.1:8000/create/'
        payload = {
            'emp_id': emp_id,
            'name': name,
            'month': month,
            'pf_status': pf_status,
            'gross': gross,
            'net_days': net_days,
            'arrears': arrears,
            'shift_allowance': shift_allowance
        }

        response = requests.post(api_url, data=payload)
        if response.ok:
            # Employee created successfully
            return render(request, 'create.html')
        else:
            # Handle the API error
            error_message = 'An error occurred while creating the employee.'
            return render(request, 'error.html', {'message': error_message})

    # If it's a GET request, simply render the HTML form
    return render(request, 'create.html')




def list_html(request):
    return render(request,'list.html')

def generate_excel_html(request):
    if request.method == 'POST':
        # Get the values from the HTML form
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')

        # Make a POST request to the generate_excel API endpoint
        response = requests.post('http://127.0.0.1:8000/generate_excel/', data={
            'from_date': from_date,
            'to_date': to_date
        })

        # Check if the request was successful
        if response.status_code == 200:
            # Get the file content from the response
            file_content = response.content
            file_name = response.headers['Content-Disposition'].split('=')[1].replace('"', '')

            # Define the directory where you want to save the Excel file
            save_directory = 'C:/INDU/django_project/django rest framework project/'

            # Create the file path by joining the directory and file name
            file_path = os.path.join(save_directory, file_name)

            # Write the file content to the specified file path
            with open(file_path, 'wb') as file:
                file.write(file_content)

            # TODO: Process the response as needed
            # For example, you can save the Excel file or display a success message
            return HttpResponse(f'Excel file generated successfully. File name: {file_name}, File path: {file_path}')
        else:
            # Handle the case where the request was not successful
            # You can display an error message or redirect to an error page
            return HttpResponse('Error: Unable to generate the Excel file.')
    else:
        # Render the HTML template
        return render(request, 'template.html')


@api_view(['POST'])
def generate_excel(request):
    print(request.data, request.method, "###############################")
    if request.method == 'POST':
        from_date = request.data.get('from_date')
        to_date = request.data.get('to_date')

        # Retrieve data from the database
        data = CallReportMaster.objects.filter(date__range=(from_date, to_date)).values(
            'sno', 'emp_id', 'ref_type', 'unique_id', 'name', 'design', 'contact', 'camp', 'camp_details',
            'date', 'time', 'location', 'latitude', 'longitude', 'area', 'city', 'state', 'pincode',
            'district', 'station', 'branch', 'source', 'attendance', 'reason', 'type', 'ldate', 'category', 'status'
        )
        
        # Create a new workbook and select the active sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Set column names
        column_names = [
            'Sno', 'Emp ID', 'Ref Type', 'Unique ID', 'Name', 'Design', 'Contact', 'Camp', 'Camp Details', 'Date',
            'Time', 'Location', 'Latitude', 'Longitude', 'Area', 'City', 'State', 'Pincode', 'District', 'Station',
            'Branch', 'Source', 'Attendance', 'Reason', 'Type', 'Ldate', 'Category', 'Status',
        ]

        # Write column names to the first row of the sheet
        for col_num, column_name in enumerate(column_names, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sheet[f'{col_letter}1'] = column_name

        # Adjust the width of each column
        for col_num in range(1, len(column_names) + 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = 12

        # Write data to the sheet
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data.values(), 1):
                col_letter = openpyxl.utils.get_column_letter(col_num)
                sheet[f'{col_letter}{row_num}'] = cell_value

        # Generate the file name
        file_name = f"report_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

        # Save the workbook to a BytesIO buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Create a FileResponse with the buffer data
        response = FileResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{file_name}"'

        return response

    
    
    return Response({'status': 400})

class employeehyperviewset(viewsets.ModelViewSet):
    queryset=Employee.objects.all()
    serializer_class=employeehyperserializer

#model viewset based one class we do all crud operations 
class employeemodalviewset(viewsets.ModelViewSet):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer
    # authentication_classes=[BasicAuthentication]
    # permission_classes=[IsAuthenticated]
    authentication_classes=[SessionAuthentication]
    permission_classes=[IsAdminUser]


#viewset
class employeeviewset(viewsets.ViewSet):
    def list(self,request):
        employee_obj=Employee.objects.all()
        serializers=EmployeeSerializer(employee_obj,many=True)
        return Response({'status':200,"payload":serializers.data})
    def retrieve(self,request,pk=None):
        id=pk
        if id is not None:
            employee_obj=Employee.objects.get(id=id)
            serializers=EmployeeSerializer(employee_obj)
            return Response({'status':200,"payload":serializers.data})
        
    def update(self,request,pk):
        id=pk
        if id is not None:
            employee_obj=Employee.objects.get(id=id)
            serializers=EmployeeSerializer(instance=employee_obj,data=request.data)
            if serializers.is_valid():
                serializers.save()
                return Response({'status':200,"payload":serializers.data})
            else:
                return Response({'status':400,"payload":serializers.errors})
    def destroy(self,request,pk):
        id=pk
        if id is not None:
            employee_obj=Employee.objects.get(id=id)
            employee_obj.delete()
            return Response("Employee is deleted successfully")
        
    def create(self,request):
        serializers=EmployeeSerializer(data=request.data)
        if serializers.is_valid():
            serializers.save()
            return Response({'status':200,"payload":serializers.data})
        else:
            return Response({'status':400,"payload":serializers.errors})


#APIView decorator
@api_view(['GET'])
def home(request):
    employee_obj=Employee.objects.all()
    serializers=EmployeeSerializer(employee_obj,many=True)
    return Response({'status':200,"payload":serializers.data})

@api_view(['POST'])
def employe_post(request):
    employee_obj=Employee.objects.all()
    serializers=EmployeeSerializer(data=request.data)
    if serializers.is_valid():
        serializers.save()
        return Response({'status':"Success","payload":serializers.data})
    else:
        return Response({'status':"Bad Request","payload":serializers.errors})    
@api_view(['POST'])
def employe_update(request,id):
    employee_obj=Employee.objects.get(id=id)
    print(employee_obj,'2222222222')
    serializers=EmployeeSerializer(instance=employee_obj, data=request.data)
    print(serializers,'#222222###')
    if serializers.is_valid():
        serializers.save()
        return Response({'status':200,"payload":serializers.data})
    else:
        return Response({'status':400,"payload":serializers.errors})      
@api_view(['DELETE'])  
def employe_delete(request,id):
    employee_obj=Employee.objects.get(id=id)
    print(employee_obj,'2222222222')
    employee_obj.delete()
    return Response("Employee is deleted successfully")        

# Mixin GenericAPIView
class Employee_list(ListModelMixin,GenericAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer
    def get(self,request):
        return self.list(request)
    

class Employee_create(CreateModelMixin,GenericAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer
    def post(self,request):
        print(type(request.data))
        print(request.data,'@@@@@@@@@@@')
        return self.create(request) 

class Employee_retrieve(RetrieveModelMixin,GenericAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer
    def get(self,request,**kwargs):
        return self.retrieve(request,**kwargs)     
      
class Employee_update(UpdateModelMixin,GenericAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer
    def put(self,request,**kwargs):
        return self.update(request,**kwargs)  

class Employee_delete(DestroyModelMixin,GenericAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer
    def delete(self,request,**kwargs):
        return self.destroy(request,**kwargs)


#concrete generic view
class employeelist(ListAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer


class employeecreate(CreateAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer

class employeeretrieve(RetrieveAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer   

class employeeupdate(UpdateAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer  

class employeedelete(DestroyAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer 

class employeelistcreate(ListCreateAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer    

class employeeretrieveupdate(RetrieveUpdateAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer

class employeeretrievedelete(RetrieveDestroyAPIView):   
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer

class employeeretrieveupdatedestroy(RetrieveUpdateDestroyAPIView):
    queryset=Employee.objects.all()
    serializer_class=EmployeeSerializer